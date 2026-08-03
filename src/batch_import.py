"""
Batch import module for CSV/Excel journal evaluation.
"""
import csv
import io
from typing import List, Dict, Any
from pathlib import Path

from src.models import JournalInput
from src.evaluator import MTUJournalEvaluator
from src.database import EvaluationDatabase


class BatchImporter:
    REQUIRED_COLUMNS = {"name", "url"}

    OPTIONAL_COLUMNS = {
        "issn_print", "issn_online", "doi_prefix", "publisher_name",
        "publisher_url", "publisher_address", "editorial_board_url",
        "submission_portal_url", "ethics_policy_url", "open_access",
        "submission_email_only", "claimed_indexes", "metric_claims",
        "rapid_publication_claim", "lock_pdfs",
    }

    @staticmethod
    def parse_csv(content: str, delimiter: str = ",") -> List[Dict[str, str]]:
        reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
        rows = []
        for row in reader:
            rows.append({k.strip(): v.strip() for k, v in row.items() if k and v is not None})
        return rows

    @staticmethod
    def parse_excel(path: str) -> List[Dict[str, str]]:
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for Excel import. Install it with: pip install openpyxl")
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        rows = []
        headers = [cell.value for cell in ws[1] if cell.value]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            rows.append({str(h).strip(): str(v).strip() if v is not None else "" for h, v in zip(headers, row)})
        wb.close()
        return rows

    @staticmethod
    def row_to_journal(row: Dict[str, str]) -> JournalInput:
        claimed_indexes = row.get("claimed_indexes", "")
        metric_claims = row.get("metric_claims", "")
        return JournalInput(
            name=row.get("name", ""),
            url=row.get("url", ""),
            issn_print=row.get("issn_print") or None,
            issn_online=row.get("issn_online") or None,
            doi_prefix=row.get("doi_prefix") or None,
            publisher_name=row.get("publisher_name") or None,
            publisher_url=row.get("publisher_url") or None,
            publisher_address=row.get("publisher_address") or None,
            editorial_board_url=row.get("editorial_board_url") or None,
            submission_portal_url=row.get("submission_portal_url") or None,
            ethics_policy_url=row.get("ethics_policy_url") or None,
            open_access=row.get("open_access", "").lower() in ("true", "1", "yes"),
            submission_email_only=row.get("submission_email_only", "").lower() in ("true", "1", "yes"),
            claimed_indexes=[x.strip() for x in claimed_indexes.split(",") if x.strip()],
            metric_claims=[x.strip() for x in metric_claims.split(",") if x.strip()],
            rapid_publication_claim=row.get("rapid_publication_claim", "").lower() in ("true", "1", "yes"),
            lock_pdfs=row.get("lock_pdfs", "").lower() in ("true", "1", "yes"),
        )

    @staticmethod
    def validate_row(row: Dict[str, str]) -> List[str]:
        errors = []
        if not row.get("name"):
            errors.append("Missing 'name'")
        if not row.get("url"):
            errors.append("Missing 'url'")
        return errors

    @staticmethod
    def batch_evaluate(rows: List[Dict[str, str]], save: bool = True) -> List[Dict[str, Any]]:
        evaluator = MTUJournalEvaluator()
        db = EvaluationDatabase() if save else None
        results = []
        for row in rows:
            errors = BatchImporter.validate_row(row)
            if errors:
                results.append({
                    "journal_name": row.get("name", "UNKNOWN"),
                    "status": "ERROR",
                    "errors": errors,
                })
                continue
            try:
                journal = BatchImporter.row_to_journal(row)
                report = evaluator.evaluate_and_report(journal, fmt="json")
                result = __import__('json').loads(report)
                if save and db:
                    db.save_evaluation(result, evaluated_by="batch_import")
                results.append(result)
            except Exception as e:
                results.append({
                    "journal_name": row.get("name", "UNKNOWN"),
                    "status": "ERROR",
                    "errors": [str(e)],
                })
        return results
